from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from rest_framework.parsers import JSONParser
from django.http.response import JsonResponse
from io import BytesIO
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Beneficiario
from django.http import HttpResponse
from openpyxl import Workbook
from django.views.generic import TemplateView

from beneficiario.models import Beneficiario, Area, Municipio
from beneficiario.serializers import BeneficiarioSerializer

@csrf_exempt
# Función para pruebas con postman
def beneficiariosApi(request, id=0):
    if request.method == 'GET':
        beneficiarios = Beneficiario.objects.all()
        beneficiarios_serializer = BeneficiarioSerializer(beneficiarios, many=True)
        return JsonResponse(beneficiarios_serializer.data, safe=False)
    elif request.method == 'POST':
        data = JSONParser().parse(request)
        beneficiarios_serializer = BeneficiarioSerializer(data = data)
        if beneficiarios_serializer.is_valid():
            beneficiarios_serializer.save()
            beneficiarios = Beneficiario.objects.all()
            return render(request, 'lista_beneficiarios.html', {'beneficiarios': beneficiarios})
        return JsonResponse(beneficiarios_serializer.errors, safe=False)
    elif request.method == 'PUT':
        data = JSONParser().parse(BytesIO(request.body))
        beneficiario = Beneficiario.objects.get(id=data["id"])
        beneficiarios_serializer = BeneficiarioSerializer(beneficiario, data=data)
        if beneficiarios_serializer.is_valid():
            beneficiarios_serializer.save()
            beneficiarios = Beneficiario.objects.all()
            return render(request, 'lista_beneficiarios.html', {'beneficiarios': beneficiarios})
        return JsonResponse(beneficiarios_serializer.data, safe=False)
    elif request.method=='DELETE':
        beneficiario = Beneficiario.objects.get(id=id)
        beneficiario.delete()
        beneficiarios = Beneficiario.objects.all()
        return render(request, 'lista_beneficiarios.html', {'beneficiarios': beneficiarios})
    
# Vistas del formulario

def formulario_beneficiario(request):
    areas_list = Area.objects.all()
    
    if request.method == 'POST':
        data = request.POST
        beneficiarios_serializer = BeneficiarioSerializer(data=data)
        
        if beneficiarios_serializer.is_valid():
            beneficiarios_serializer.save()
            return redirect('lista_beneficiarios')
        
        return render(request, 'formulario_registro.html', {
            'serializer': beneficiarios_serializer,
            'areas_list': areas_list
        })

    beneficiarios_serializer = BeneficiarioSerializer()
    return render(request, 'formulario_registro.html', {
        'serializer': beneficiarios_serializer,
        'areas_list': areas_list
    })
    
def agregar_beneficiario(request):
    if request.method == 'POST':
        data = request.POST
        beneficiarios_serializer = BeneficiarioSerializer(data = data)
        if beneficiarios_serializer.is_valid():
            beneficiarios_serializer.save()
            return redirect('lista_beneficiarios')
        else:
            return JsonResponse(beneficiarios_serializer.errors, status=400)
    else:
        return render(request, 'formulario_registro.html')

def lista_beneficiarios(request):
    if request.method == 'POST':
        area_query = request.POST.get('area_search', '')
        ejercicio_query = request.POST.get('ejercicio_search', '')
        municipio_query = request.POST.get('municipio_search', '')
        page_number = request.POST.get('page', 1)  # Obtiene el número de página enviado desde el formulario
    else:
        area_query = ''
        ejercicio_query = ''
        municipio_query = ''
        page_number = 1  # Por defecto a la primera página

    beneficiarios_list = Beneficiario.objects.all()
    municipios_list = Municipio.objects.all()
    areas_list = Area.objects.all()
    consulta_realizada = False
    
    beneficiarios_unicos = beneficiarios_list.values_list("folio_sd", flat=True).distinct()
    total_beneficiarios_unicos = len(beneficiarios_unicos)
    
    programa_unico = beneficiarios_list.values_list("programa", flat=True).distinct()
    total_programa_unico = len(programa_unico)
    
    if area_query or ejercicio_query or municipio_query:
        beneficiarios_list = beneficiarios_list.filter(
            Q(area_programa__icontains=area_query) &
            Q(ejercicio__icontains=ejercicio_query) &
            Q(municipio__icontains=municipio_query)
        )
        
        total_beneficiarios_busqueda = beneficiarios_list.count()
        mostrarResultados = True
        consulta_realizada = True  # Marcar como consulta realizada
        
        # Implementar paginación
        paginator = Paginator(beneficiarios_list, 10)
        beneficiarios = paginator.get_page(page_number)

        return render(request, 'lista_beneficiarios.html', {
            'beneficiarios': beneficiarios, 
            'municipios': municipios_list, 
            'areas': areas_list,
            'area_query': area_query, 
            'ejercicio_query': ejercicio_query, 
            'municipio_query': municipio_query,
            'total_beneficiarios': total_beneficiarios_busqueda, 
            'mostrarResultados': mostrarResultados,
            'total_beneficiarios_unicos': total_beneficiarios_unicos,
            'total_programa_unico': total_programa_unico,
            'consulta_realizada': consulta_realizada
        })

    else:
        beneficiarios_list = [] 
        total_beneficiarios_busqueda = 0
        mostrarResultados = False

        return render(request, 'lista_beneficiarios.html', {
            'municipios': municipios_list, 
            'areas': areas_list,
            'total_beneficiarios': total_beneficiarios_busqueda, 
            'mostrarResultados': mostrarResultados,
            'total_beneficiarios_unicos': total_beneficiarios_unicos,
            'total_programa_unico': total_programa_unico,
            'consulta_realizada': consulta_realizada
        })

class ReportebenExcel(TemplateView):
    def get(self, *args, **kwargs):
        beneficiarios = Beneficiario.objects.all()
        wb = Workbook()
        ws = wb.active

        ws['B1'] = 'REPORTE'
        ws.merge_cells('B1:I1')
        ws['B3'] = 'NOMBRE'
        ws['C3'] = 'APELLIDO PATERNO'
        ws['D3'] = 'APELLIDO MATERNO'
        ws['E3'] = 'SEXO'
        ws['F3'] = 'PROGRAMA'
        ws['G3'] = 'AREA'
        ws['H3'] = 'EJERCICIO'
        ws['I3'] = 'MUNICIPIO'

        cont = 4  # Ajustado a 4 para que inicie después del encabezado

        for beneficiario in beneficiarios:
            ws.cell(row=cont, column=2).value = beneficiario.nombres
            ws.cell(row=cont, column=3).value = beneficiario.apellido_paterno
            ws.cell(row=cont, column=4).value = beneficiario.apellido_materno
            ws.cell(row=cont, column=5).value = beneficiario.sexo
            ws.cell(row=cont, column=6).value = beneficiario.programa
            ws.cell(row=cont, column=7).value = beneficiario.area_programa
            ws.cell(row=cont, column=8).value = beneficiario.ejercicio
            ws.cell(row=cont, column=9).value = beneficiario.municipio
            cont += 1

        # Configuración de la respuesta HTTP con el archivo
        nombre_archivo = "Reporte Padron de Beneficiarios.xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        
        # Guardar el archivo en la respuesta
        wb.save(response)
        return response
    
def editar_beneficiario(request, id):
    beneficiario = Beneficiario.objects.get(id=id)
    if request.method == 'POST':
        beneficiario.nombres = request.POST['nombres']
        beneficiario.save()
        return redirect('lista_beneficiarios')
    else:
        return render(request, 'formulario_editar.html', {'beneficiario': beneficiario})

def eliminar_beneficiario(request, id):
    beneficiario = Beneficiario.objects.get(id=id)
    beneficiario.delete()
    return redirect('lista_beneficiarios')